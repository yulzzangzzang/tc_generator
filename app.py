import streamlit as st
import os
import pandas as pd
import fitz  # PyMuPDF
from io import BytesIO
from datetime import datetime
from openai import OpenAI
import time
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# --- [1. 설정 및 API 키] ---
# OpenAI API 키를 입력하세요. (sk-...)
API_KEY = "OPENAI_API_KEY"
client = OpenAI(api_key=API_KEY)


# --- [2. PDF/Excel 읽기 함수] ---
def get_pdf_text_from_upload(uploaded_files):
    all_text = ""
    for uploaded_file in uploaded_files:
        try:
            doc = fitz.open(stream=uploaded_file.read(), filetype="pdf")
            for page in doc:
                all_text += page.get_text()
            doc.close()
        except Exception as e:
            st.error(f"❌ [PDF 오류] {uploaded_file.name}: {e}")
    return all_text


def get_old_excel_data(uploaded_excel):
    if uploaded_excel:
        try:
            df = pd.read_excel(uploaded_excel)
            return df.to_string(index=False)
        except:
            return None
    return None


# --- [3. 스타일링 함수 (미리보기용)] ---
def highlight_tc_rows(row):
    note = str(row.비고)
    if '[수정]' in note:
        return ['background-color: #FFFF00'] * len(row)
    elif '[신규]' in note:
        return ['background-color: #CCEEFF'] * len(row)
    elif '[삭제]' in note:
        return ['background-color: #D3D3D3'] * len(row)
    return [''] * len(row)


# --- [4. 웹 화면 구성] ---
st.set_page_config(page_title="QA TC Generator Pro", layout="wide")
st.title("🚀 테스트 케이스 생성기 (OpenAI)")

st.info("💡 **처음 제작**: 기획서만 업로드 / **수정 업데이트**: 이전 엑셀을 추가로 업로드")

col1, col2 = st.columns(2)
with col1:
    st.subheader("📁 1. 기획서 업로드 (필수)")
    uploaded_files = st.file_uploader("기획서 PDF (여러 개 가능)", type="pdf", accept_multiple_files=True)
with col2:
    st.subheader("📂 2. 기존 TC 업로드 (선택)")
    old_excel = st.file_uploader("이전에 다운받은 엑셀 파일", type="xlsx")

if uploaded_files:
    is_update = old_excel is not None
    button_label = "🪄 변경 사항 분석 및 업데이트" if is_update else "🪄 테스트 케이스 신규 생성"

    if st.button(button_label, type="primary"):
        with st.spinner("기획서 분석 및 데이터 정밀 추출 중..."):

            plan_content = get_pdf_text_from_upload(uploaded_files)
            old_data_text = get_old_excel_data(old_excel)

            # 모드별 특화 지시문
            mode_instruction = ""
            if is_update:
                mode_instruction = f"""
                ### [업데이트 모드]
                1. 제공된 [기존 데이터]를 새로운 기획서와 정밀하게 비교하라.
                2. 바뀐 부분은 '비고'에 [수정], 새로 생긴 건 [신규], 없어진 건 [삭제 대상]이라 표기하라.
                3. 기존 'TC ID' 및 구조를 유지하며 업데이트하라.
                """
            else:
                mode_instruction = f"""
                ### [신규 생성 모드]
                1. 기획서 내용을 바탕으로 화면 요구사항을 처음부터 꼼꼼하게 추출하라.
                """

            # 🚨 [사용자 요청: 기존 마스터 프롬프트 100% 보존]
            prompt = f"""
            너는 QA 엔지니어이며 TC 작성 전문가이다.
            기획서에 작성된 UI 요소 및 Description에 따라 TC를 작성해라.
            출력은 반드시 '|'로 구분된 13개 컬럼 표 형식이어야 한다.

            {mode_instruction}

            ### [핵심 미션]
            - 기획서에 명시된 모든 UI 요소(아이콘 / 버튼 / 인풋박스 / 필터 등)를 빠짐없이 도출하라.

            ### [ISTQB 기반 테스트 설계 규칙]
            1. **경계값 분석 (Boundary Value Analysis)**: 
            - 입력란(숫자, 글자 수 등)에 제한이 있는 경우, [최솟값-1, 최솟값, 최솟값+1, 최댓값-1, 최댓값, 최댓값+1] 등 경계값을 확인하는 케이스를 반드시 포함한다.
            2. **동등 분할 (Equivalence Partitioning)**: 
            - 유효한 입력 값(Pass)뿐만 아니라 유효하지 않은 입력 값(Fail) 군집을 정의하여 각각 최소 1개 이상의 케이스를 작성한다.
            3. **에러 추측 (Error Guessing)**: 
            - 기획서에 명시되지 않았더라도 '특수문자 입력', '공백 입력', '중복 클릭', '뒤로가기 시 데이터 유지' 등 시니어 QA로서 예상되는 결함 시나리오를 추가한다.
            4. **결정 테이블 (Decision Table)**: 
            - 여러 조건이 복합적으로 얽힌 로직(예: 권한별 접근 제어, 조건별 할인 등)은 조건의 조합에 따른 결과 값을 각각 별개의 행으로 작성한다.

            ### [TC 구성 및 위계]
            1. 화면 진입 및 전체 레이아웃 확인 케이스를 최상단에 배치하라.
            2. **Label 위계**: 
               - Label 1: 대분류 영역 명칭
                    - 레이아웃 확인 케이스에서 나열한 항목명을 그대로 사용하라.(예 : 로고 영역 / 아이디 영역 / 비밀번호 영역 / 검색 영역 / 로그인 버튼)
                    - Label 1의 첫 번째 행은 반드시 해당 영역의 전체 구성을 확인하는 케이스여야 한다.
               - Label 2: 구체적 확인 대상 **영역별 첫 번째 케이스 (Layout Check)**:
                    - Label 1에서 나열된 항목의 세부 테스트 케이스를 작성하며, Label 2 : "UI 확인", Label 3: "-", 수행 절차: "검색 영역의 구성을 확인한다.", 기대결과는 "항목명 / 검색 인풋박스 / 검색 버튼으로 구성되어 노출된다.")
                    - 구체적인 컴포넌트 명칭 사용(예: 아이디 인풋박스, 로그인 버튼)
                    - 다음 케이스 부터는 앞서 나열한 항목의 UI 확인 / 기능 확인 순으로 작성된다.
               - Label 3: 확인 성격 (UI 확인 / 기능 확인 / 밸리데이션 확인)
                    - Label 2와 Label 3의 명칭이 동일한 경우, Label 3는 반드시 '-'로 표기한다.
                    - Label 3에 작성될 항목이나 기능명이 없을 경우 '-로 표기한다.
                    **구성 요소별 단계적 시나리오 (Sequential Order)**:
                        아래의 순서대로 **반드시 행을 생성**하라.
                            ① **UI 확인**: 컴포넌트의 노출 상태, 레이블, Placeholder 등을 확인.
                            ② **기능 확인**: 클릭, 입력 값 표시, 마스킹 처리 등 기본 동작 확인.
                            ③ **밸리데이션 확인**: 경계값(29/30/31자), 특수문자, 공백 등 유효성 체크.
            4. **문구 규칙**: 사전 조건 / 참고 컬럼은 명사형 간결체로 작성하고 수행 절차와 내용이 중복되지 않게 하라. (예: 30자 미만 입력(ex.abcdefg), 미입력 상태)
                - 해당 행에서 테스트할 **단 하나의 구체적인 입력 조건 또는 데이터**만 명시하라.

            ### [TC 작성 문체 통일]
            1. **수행 절차 (간결화)**:
            - 사용자 스타일인 **"조건에 맞게 입력한다."** 또는 **"항목별 노출 여부를 확인한다."**라는 문구로 통일하되, 각 행의 맥락에 맞게 작성하라.
            - 예시 :
                - **로그인 관련 영역**: "조건에 맞게 입력한 뒤 [로그인] 버튼을 클릭한다."
                - **검색 관련 영역**: "조건에 맞게 입력한 뒤 [검색] 버튼을 클릭한다."
                - **일반 입력/선택**: "조건에 맞게 입력/선택 후 결과를 확인한다."
                - **UI 확인**: "해당 영역의 구성 요소 및 노출 상태를 확인한다."

            2. **기대 결과 (단일 결과)**:
            - 해당 행의 사전 조건에 명시된 특정 데이터에 대한 **단 하나의 예상 결과**만 정확히 기술하라.
            - 예시: 29자가 정상적으로 입력되고 표시된다. / 30자까지만 입력되고 31자부터는 입력되지 않는다.

            3. 구분자
            / 기호로 항목의 구분을 한다.

            ### [작성 예시]
            | TC ID | 프로그램명(화면명) | 화면 ID | 요구사항 ID | Label 1 | Label 2 | Label 3 | 사전 조건 / 참고 | 수행 절차 | 기대 결과 | 결과 | 수행자 | 비고 |
            |---|---|---|---|---|---|---|---|---|---|---|---|---|
            | | 로그인 | - | - | 대분류 영역 명칭 | 구체적 확인 대상 | 확인 성격 | 아이디 미입력 | 아이디 입력 영역을 확인한다. | 가이드 문구가 노출된다. | | | |

            [기존 데이터]
            {old_data_text if old_data_text else "없음"}

            [기획서 내용]
            {plan_content}
            """

            try:
                # OpenAI API 호출 (gpt-4o 모델 사용)
                response = client.chat.completions.create(
                    model="gpt-4o",
                    messages=[{"role": "system", "content": "You are a professional QA engineer expert."},
                              {"role": "user", "content": prompt}]
                )

                raw_data = response.choices[0].message.content.strip()
                lines = [line.strip() for line in raw_data.split('\n') if '|' in line]
                lines = [line for line in lines if not all(c in '| -:' for c in line)]

                table_data = []
                for line in lines:
                    cells = [cell.strip() for cell in line.split('|')]
                    if cells and cells[0] == "": cells.pop(0)
                    if cells and cells[-1] == "": cells.pop()
                    if len(cells) >= 10 and "TC ID" not in cells[0]:
                        if len(cells) < 13: cells.extend([""] * (13 - len(cells)))
                        table_data.append(cells[:13])

                if table_data:
                    columns = ["TC ID", "프로그램명(화면명)", "화면 ID", "요구사항 ID", "Label 1", "Label 2", "Label 3", "사전 조건 / 참고",
                               "수행 절차", "기대 결과", "결과", "수행자", "비고"]
                    df = pd.DataFrame(table_data, columns=columns)

                    df.loc[df['Label 2'] == df['Label 3'], 'Label 3'] = '-'
                    if not is_update:
                        df['TC ID'] = "";
                        df['결과'] = "Not Tested";
                        df['수행자'] = ""

                    # --- 엑셀 스타일링 로직 ---
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='Test Case')
                        ws = writer.sheets['Test Case']

                        yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                        blue = PatternFill(start_color='CCEEFF', end_color='CCEEFF', fill_type='solid')
                        gray = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                        header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                        header_font = Font(name='맑은 고딕', size=9, bold=True)
                        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                             bottom=Side(style='thin'))

                        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=len(df) + 1), 1):
                            note = str(ws.cell(row=r_idx, column=13).value)
                            for c_idx, cell in enumerate(row, 1):
                                cell.border = thin_border
                                if r_idx == 1:
                                    cell.fill = header_fill;
                                    cell.font = header_font
                                    cell.alignment = Alignment(horizontal='center', vertical='center')
                                else:
                                    cell.font = Font(name='맑은 고딕', size=9)
                                    if "[수정]" in note:
                                        cell.fill = yellow
                                    elif "[신규]" in note:
                                        cell.fill = blue
                                    elif "[삭제]" in note:
                                        cell.fill = gray
                                    align = 'left' if c_idx in [9, 10, 13] else 'center'
                                    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)

                        column_widths = [10, 15, 12, 12, 15, 18, 15, 18, 35, 35, 10, 10, 25]
                        for i, width in enumerate(column_widths, 1):
                            ws.column_dimensions[chr(64 + i)].width = width

                    st.balloons()
                    st.success(f"✅ 분석 완료! 총 {len(df)}개의 케이스가 생성되었습니다.")
                    st.dataframe(df.style.apply(highlight_tc_rows, axis=1), use_container_width=True)

                    st.download_button(
                        label="📥 테스트 케이스 다운로드 (Excel)",
                        data=output.getvalue(),
                        file_name=f"TC_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            except Exception as e:
                st.error(f"❌ OpenAI 에러: {e}")